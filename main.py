# ============================================================
# main.py — Основной файл запуска Flask-приложения
# ============================================================

from dotenv import load_dotenv

load_dotenv()

from datetime import timedelta

from flask import (Flask, render_template, request, redirect, url_for, flash,
                   jsonify, send_file, session, abort)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from flask_wtf.csrf import CSRFProtect, CSRFError
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from functools import wraps

import config
from config import (
    SECRET_KEY,
    DEBUG,
    POSITIONS,
    SEMESTERS,
    DEPARTMENT_NAME,
    PROJECT_THEME,
    PROJECT_DESCRIPTION,
    ACADEMIC_YEAR_DEFAULT,
    GROQ_MODEL,
    HOURS_PER_CREDIT,
    SUPERVISOR_ONLY_POSITIONS,
    RATES,
    DEFAULT_RATE,
    LANGUAGES,
    DEFAULT_LANG,
    LOGIN_RATE_LIMIT,
    PERMANENT_SESSION_LIFETIME_DAYS,
)
from i18n import t as _t, normalize_lang, position_label, role_label
from database import init_db
from security import install_security_headers, safe_next_url, client_ip
import teachers as tch
import subjects as sub
import workload as wld
import reports as rep
import auth as au
from ai_engine import ai
from utils import workload_status


def _parse_credits_form(form) -> float:
    raw = (form.get("credits") or "0").strip().replace(",", ".")
    if raw == "":
        return 0.0
    try:
        c = float(raw)
    except ValueError as e:
        raise ValueError(_t("error.credits_number", get_current_lang())) from e
    if c < 0 or c > 40:
        raise ValueError(_t("error.credits_range", get_current_lang()))
    return round(c, 2)


def _parse_credits_field(form, field: str, max_value: float = 40.0) -> float:
    """Парсит поле формы как кредиты (≥ 0). Запятая допускается."""
    raw = (form.get(field) or "0").strip().replace(",", ".")
    if raw == "":
        return 0.0
    try:
        c = float(raw)
    except ValueError as e:
        raise ValueError(_t("error.field_number", get_current_lang(), field=field)) from e
    if c < 0 or c > max_value:
        raise ValueError(
            _t("error.field_range", get_current_lang(), field=field, max=max_value)
        )
    return round(c, 2)


app = Flask(__name__)

# --- SECRET_KEY: обязателен в проде ---
if not SECRET_KEY:
    raise RuntimeError(
        "SECRET_KEY не задан. Установите переменную окружения SECRET_KEY "
        "(например, через .env) или запустите с FLASK_DEBUG=1 для локальной разработки."
    )
app.secret_key = SECRET_KEY

# --- Cookie / session hardening ---
app.config.update(
    SESSION_COOKIE_HTTPONLY=config.SESSION_COOKIE_HTTPONLY,
    SESSION_COOKIE_SAMESITE=config.SESSION_COOKIE_SAMESITE,
    SESSION_COOKIE_SECURE=config.SESSION_COOKIE_SECURE,
    REMEMBER_COOKIE_HTTPONLY=config.REMEMBER_COOKIE_HTTPONLY,
    REMEMBER_COOKIE_SAMESITE=config.REMEMBER_COOKIE_SAMESITE,
    REMEMBER_COOKIE_SECURE=config.REMEMBER_COOKIE_SECURE,
    PERMANENT_SESSION_LIFETIME=timedelta(days=PERMANENT_SESSION_LIFETIME_DAYS),
    # Ограничим размер тела запроса (защита от простых DoS через огромные тела)
    MAX_CONTENT_LENGTH=int(16 * 1024 * 1024),  # 16 MiB
    WTF_CSRF_TIME_LIMIT=int(60 * 60 * 4),      # 4 часа на CSRF-токен
)

# --- CSRF protection (все POST/PUT/DELETE формы) ---
csrf = CSRFProtect(app)

# --- Rate limiting (brute-force защита) ---
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://",
)

# --- Security headers на каждый ответ ---
install_security_headers(app)


@app.errorhandler(CSRFError)
def _handle_csrf_error(e):
    flash(_t("csrf.invalid", get_current_lang()), "danger")
    return redirect(request.referrer or url_for("index"))


def get_current_lang() -> str:
    """Возвращает текущий язык интерфейса (из сессии или по умолчанию)."""
    return normalize_lang(session.get("lang", DEFAULT_LANG))


@app.context_processor
def inject_template_globals():
    lang = get_current_lang()

    def t(key, **kwargs):
        return _t(key, lang, **kwargs)

    def pos_label(p):
        return position_label(p, lang)

    def rl_label(r):
        return role_label(r, lang)

    return {
        "dept": _t("project.department", lang),
        "project_theme": _t("project.theme", lang),
        "project_description": _t("project.description", lang),
        "academic_year_default": ACADEMIC_YEAR_DEFAULT,
        "hours_per_credit": HOURS_PER_CREDIT,
        "t": t,
        "position_label": pos_label,
        "role_label": rl_label,
        "current_lang": lang,
        "languages": LANGUAGES,
    }


@app.route("/set-language/<lang>")
def set_language(lang):
    """Переключатель языка интерфейса (RU / KZ)."""
    session["lang"] = normalize_lang(lang)
    session.permanent = True
    return redirect(request.referrer or url_for("index"))


# ── Flask-Login ──────────────────────────────────────────────
login_manager = LoginManager(app)
login_manager.login_view             = "login"
login_manager.login_message_category = "warning"


@login_manager.unauthorized_handler
def _unauthorized():
    flash(_t("login.please_login", get_current_lang()), "warning")
    return redirect(url_for("login", next=request.path))

@login_manager.user_loader
def load_user(user_id):
    return au.get_user_by_id(int(user_id))

# ── Инициализация БД при старте ─────────────────────────────
with app.app_context():
    init_db()
    au.ensure_default_admin()


# ── Декоратор: только admin ──────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash(_t("error.access.admin", get_current_lang()), "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


# ── Декоратор: admin или editor ──────────────────────────────
def editor_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.can_edit():
            flash(_t("error.access.editor", get_current_lang()), "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


# ════════════════════════════════════════════════════════════
# ВХОД / ВЫХОД
# ════════════════════════════════════════════════════════════
@app.route("/login", methods=["GET", "POST"])
@limiter.limit(
    LOGIN_RATE_LIMIT,
    methods=["POST"],
    error_message="rate_limited",
)
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        remember = bool(request.form.get("remember"))
        ip = client_ip()
        ua = request.headers.get("User-Agent", "")
        user = au.verify_password(username, password)
        au.record_login_attempt(username, ip, success=bool(user), user_agent=ua)
        if user:
            # Защита от session fixation: сбрасываем текущую сессию
            # (сохраняя только выбранный язык), а затем логиним пользователя.
            saved_lang = session.get("lang")
            session.clear()
            if saved_lang:
                session["lang"] = saved_lang
            login_user(user, remember=remember)
            session.permanent = True
            flash(_t("login.welcome", get_current_lang(), name=user.full_name), "success")
            if user.must_change_password:
                flash(_t("login.must_change_password", get_current_lang()), "warning")
                return redirect(url_for("profile"))
            return redirect(safe_next_url(request.args.get("next"), url_for("index")))
        flash(_t("login.bad_credentials", get_current_lang()), "danger")
    return render_template("login.html")


@app.errorhandler(429)
def _handle_rate_limit(e):
    # Единое сообщение для превышения лимита (в т.ч. /login).
    if request.path.startswith("/login"):
        flash(_t("login.rate_limited", get_current_lang()), "danger")
        return redirect(url_for("login"))
    return ("Too Many Requests", 429)


@app.route("/logout", methods=["POST"])
@login_required
def logout():
    # Полностью очищаем сессию — не оставляем чувствительных данных
    saved_lang = session.get("lang")
    logout_user()
    session.clear()
    if saved_lang:
        session["lang"] = saved_lang
    flash(_t("login.logged_out", get_current_lang()), "info")
    return redirect(url_for("login"))


@app.before_request
def _enforce_password_change():
    """Если у пользователя стоит флаг must_change_password — блокируем всё,
    кроме страницы профиля / logout / static / смены языка."""
    if not current_user.is_authenticated:
        return None
    if not getattr(current_user, "must_change_password", False):
        return None
    allowed = {"profile", "logout", "static", "set_language"}
    if request.endpoint in allowed:
        return None
    return redirect(url_for("profile"))


# ════════════════════════════════════════════════════════════
# ГЛАВНАЯ СТРАНИЦА
# ════════════════════════════════════════════════════════════
@app.route("/")
@login_required
def index():
    lang = get_current_lang()
    teachers_summary = tch.get_teachers_with_workload_summary()
    stats            = wld.get_workload_stats()
    overloaded       = rep.report_overloaded_teachers(lang)
    data = []
    for t in teachers_summary:
        row = dict(t)
        row["status"] = workload_status(row["total_credits"], row["max_workload"], lang)
        data.append(row)
    return render_template("index.html", teachers=data, stats=stats,
                           overloaded_count=len(overloaded))


# ════════════════════════════════════════════════════════════
# ПРЕПОДАВАТЕЛИ
# ════════════════════════════════════════════════════════════
@app.route("/teachers")
@login_required
def teachers_list():
    return render_template("teachers.html", teachers=tch.get_all_teachers(),
                           positions=POSITIONS)


def _parse_rate(form) -> float:
    raw = (form.get("rate") or str(DEFAULT_RATE)).strip().replace(",", ".")
    try:
        r = float(raw)
    except ValueError as e:
        raise ValueError(_t("error.rate_number", get_current_lang())) from e
    if r <= 0 or r > 2.0:
        raise ValueError(_t("error.rate_range", get_current_lang()))
    return round(r, 2)


@app.route("/teachers/add", methods=["GET", "POST"])
@login_required
@editor_required
def teacher_add():
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        position  = request.form.get("position", "")
        email     = request.form.get("email", "").strip()
        if not full_name or not position:
            flash(_t("error.required", get_current_lang()), "danger")
        else:
            try:
                rate = _parse_rate(request.form)
                tch.add_teacher(
                    full_name, position,
                    _parse_credits_field(request.form, "max_workload", max_value=200),
                    email,
                    rate,
                )
                flash(_t("teachers.flash.added", get_current_lang(), name=full_name), "success")
                return redirect(url_for("teachers_list"))
            except ValueError as e:
                flash(_t("error.form_data", get_current_lang(), msg=str(e)), "danger")
    from config import POSITION_MAX_WORKLOAD
    return render_template("teacher_form.html", action="add", teacher=None,
                           positions=POSITIONS, position_workloads=POSITION_MAX_WORKLOAD,
                           rates=RATES, default_rate=DEFAULT_RATE,
                           supervisor_only=list(SUPERVISOR_ONLY_POSITIONS))


@app.route("/teachers/edit/<int:tid>", methods=["GET", "POST"])
@login_required
@editor_required
def teacher_edit(tid):
    teacher = tch.get_teacher_by_id(tid)
    if not teacher:
        flash(_t("teachers.flash.not_found", get_current_lang()), "danger")
        return redirect(url_for("teachers_list"))
    if request.method == "POST":
        try:
            rate = _parse_rate(request.form)
            tch.update_teacher(
                tid,
                request.form.get("full_name", "").strip(),
                request.form.get("position", ""),
                _parse_credits_field(request.form, "max_workload", max_value=200),
                request.form.get("email", "").strip(),
                rate,
            )
            flash(_t("teachers.flash.updated", get_current_lang()), "success")
            return redirect(url_for("teachers_list"))
        except ValueError as e:
            flash(_t("error.form_data", get_current_lang(), msg=str(e)), "danger")
    from config import POSITION_MAX_WORKLOAD
    return render_template("teacher_form.html", action="edit", teacher=dict(teacher),
                           positions=POSITIONS, position_workloads=POSITION_MAX_WORKLOAD,
                           rates=RATES, default_rate=DEFAULT_RATE,
                           supervisor_only=list(SUPERVISOR_ONLY_POSITIONS))


@app.route("/teachers/delete/<int:tid>", methods=["POST"])
@login_required
@editor_required
def teacher_delete(tid):
    t = tch.get_teacher_by_id(tid)
    if t:
        tch.delete_teacher(tid)
        flash(_t("teachers.flash.deleted", get_current_lang(), name=t['full_name']), "warning")
    return redirect(url_for("teachers_list"))


# ════════════════════════════════════════════════════════════
# ДИСЦИПЛИНЫ
# ════════════════════════════════════════════════════════════
@app.route("/subjects")
@login_required
def subjects_list():
    return render_template("subjects.html", subjects=sub.get_all_subjects())


@app.route("/subjects/add", methods=["GET", "POST"])
@login_required
@editor_required
def subject_add():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash(_t("subjects.flash.name_required", get_current_lang()), "danger")
        else:
            try:
                cr = _parse_credits_form(request.form)
                sub.add_subject(
                    name,
                    request.form.get("code", "").strip(),
                    request.form.get("description", "").strip(),
                    cr,
                )
                flash(_t("subjects.flash.added", get_current_lang(), name=name), "success")
                return redirect(url_for("subjects_list"))
            except ValueError as e:
                flash(str(e), "danger")
    return render_template("subject_form.html", action="add", subject=None)


@app.route("/subjects/edit/<int:sid>", methods=["GET", "POST"])
@login_required
@editor_required
def subject_edit(sid):
    subject = sub.get_subject_by_id(sid)
    if not subject:
        flash(_t("subjects.flash.not_found", get_current_lang()), "danger")
        return redirect(url_for("subjects_list"))
    if request.method == "POST":
        try:
            cr = _parse_credits_form(request.form)
            sub.update_subject(
                sid,
                request.form.get("name", "").strip(),
                request.form.get("code", "").strip(),
                request.form.get("description", "").strip(),
                cr,
            )
            flash(_t("subjects.flash.updated", get_current_lang()), "success")
            return redirect(url_for("subjects_list"))
        except ValueError as e:
            flash(str(e), "danger")
    return render_template("subject_form.html", action="edit",
                           subject=dict(subject))


@app.route("/subjects/delete/<int:sid>", methods=["POST"])
@login_required
@editor_required
def subject_delete(sid):
    s = sub.get_subject_by_id(sid)
    if s:
        sub.delete_subject(sid)
        flash(_t("subjects.flash.deleted", get_current_lang(), name=s['name']), "warning")
    return redirect(url_for("subjects_list"))


# ════════════════════════════════════════════════════════════
# УЧЕБНАЯ НАГРУЗКА
# ════════════════════════════════════════════════════════════
@app.route("/workload")
@login_required
def workload_list():
    return render_template("workload.html", entries=wld.get_all_workload())


def _check_supervisor_only(teacher_id: int,
                           lec: float, prc: float, lab: float) -> None:
    """Запрещает учебную нагрузку преподавателям с ролью "только жетекші"
    (например, «Исследователь»). Поднимает ValueError, если нарушение."""
    teacher = tch.get_teacher_by_id(teacher_id)
    if not teacher:
        return
    if teacher["position"] in SUPERVISOR_ONLY_POSITIONS and (lec + prc + lab) > 0:
        raise ValueError(_t(
            "error.supervisor_only",
            get_current_lang(),
            name=teacher['full_name'],
            position=position_label(teacher['position'], get_current_lang()),
        ))


@app.route("/workload/add", methods=["GET", "POST"])
@login_required
@editor_required
def workload_add():
    if request.method == "POST":
        try:
            teacher_id = int(request.form["teacher_id"])
            lec = _parse_credits_field(request.form, "lecture_credits")
            prc = _parse_credits_field(request.form, "practice_credits")
            lab = _parse_credits_field(request.form, "lab_credits")
            _check_supervisor_only(teacher_id, lec, prc, lab)
            wld.assign_workload(
                teacher_id, int(request.form["subject_id"]),
                lec, prc, lab,
                int(request.form.get("semester", 1)),
                request.form.get("academic_year", ACADEMIC_YEAR_DEFAULT).strip())
            flash(_t("workload.flash.assigned", get_current_lang()), "success")
            return redirect(url_for("workload_list"))
        except Exception as e:
            flash(_t("workload.flash.error", get_current_lang(), msg=str(e)), "danger")
    return render_template("workload_form.html", action="add", entry=None,
                           teachers=tch.get_all_teachers(), subjects=sub.get_all_subjects(),
                           semesters=SEMESTERS,
                           supervisor_only=list(SUPERVISOR_ONLY_POSITIONS))


@app.route("/workload/edit/<int:wid>", methods=["GET", "POST"])
@login_required
@editor_required
def workload_edit(wid):
    entry = wld.get_workload_entry(wid)
    if not entry:
        flash(_t("workload.flash.not_found", get_current_lang()), "danger")
        return redirect(url_for("workload_list"))
    if request.method == "POST":
        try:
            lec = _parse_credits_field(request.form, "lecture_credits")
            prc = _parse_credits_field(request.form, "practice_credits")
            lab = _parse_credits_field(request.form, "lab_credits")
            _check_supervisor_only(entry["teacher_id"], lec, prc, lab)
            wld.update_workload(wid, lec, prc, lab,
                int(request.form.get("semester", 1)),
                request.form.get("academic_year", ACADEMIC_YEAR_DEFAULT).strip())
            flash(_t("workload.flash.updated", get_current_lang()), "success")
            return redirect(url_for("workload_list"))
        except Exception as e:
            flash(_t("workload.flash.error", get_current_lang(), msg=str(e)), "danger")
    return render_template("workload_form.html", action="edit", entry=dict(entry),
                           teachers=tch.get_all_teachers(), subjects=sub.get_all_subjects(),
                           semesters=SEMESTERS,
                           supervisor_only=list(SUPERVISOR_ONLY_POSITIONS))


@app.route("/workload/delete/<int:wid>", methods=["POST"])
@login_required
@editor_required
def workload_delete(wid):
    wld.delete_workload(wid)
    flash(_t("workload.flash.deleted", get_current_lang()), "warning")
    return redirect(url_for("workload_list"))


@app.route("/workload/export.xlsx")
@login_required
def workload_export_xlsx():
    """Выгрузка всей нагрузки в Excel-файл (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash(_t("error.excel_missing", get_current_lang()), "danger")
        return redirect(url_for("workload_list"))

    import io
    from datetime import datetime

    entries = wld.get_all_workload()

    wb = Workbook()
    ws = wb.active
    ws.title = "Нагрузка"

    headers = [
        "Преподаватель", "Должность", "Дисциплина", "Код",
        "Семестр", "Лекции, кр.", "Практика, кр.", "Лаб., кр.", "Итого, кр.",
        "Уч. год",
    ]
    ws.append(headers)

    header_font   = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    header_fill   = PatternFill("solid", fgColor="FF2F7DF9")
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin          = Side(border_style="thin", color="FFE3E8F2")
    cell_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font      = header_font
        c.fill      = header_fill
        c.alignment = center_align
        c.border    = cell_border

    for e in entries:
        ws.append([
            e["teacher_name"],
            e["position"],
            e["subject_name"],
            e["subject_code"] or "",
            e["semester"],
            e["lecture_credits"],
            e["practice_credits"],
            e["lab_credits"],
            e["total_credits"],
            e["academic_year"],
        ])

    last_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last_row,
                            min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = cell_border
            if cell.column >= 5:  # числовые/коротко-текстовые столбцы
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Итоговая строка
    if entries:
        total_lecture  = round(sum(e["lecture_credits"]  for e in entries), 2)
        total_practice = round(sum(e["practice_credits"] for e in entries), 2)
        total_lab      = round(sum(e["lab_credits"]      for e in entries), 2)
        total_all      = round(total_lecture + total_practice + total_lab, 2)
        ws.append([
            "ИТОГО", "", "", "", "",
            total_lecture, total_practice, total_lab, total_all, "",
        ])
        total_row = ws.max_row
        total_font = Font(bold=True)
        total_fill = PatternFill("solid", fgColor="FFF0F4FB")
        for cell in ws[total_row]:
            cell.font      = total_font
            cell.fill      = total_fill
            cell.border    = cell_border
            cell.alignment = center_align

    widths = [28, 22, 32, 12, 8, 10, 10, 8, 11, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"workload_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/workload/teacher/<int:tid>")
@login_required
def workload_teacher(tid):
    lang = get_current_lang()
    teacher = tch.get_teacher_by_id(tid)
    if not teacher:
        flash(_t("teachers.flash.not_found", lang), "danger")
        return redirect(url_for("workload_list"))
    entries       = wld.get_workload_by_teacher(tid)
    total_credits = wld.get_teacher_total_credits(tid)
    status        = workload_status(total_credits, teacher["max_workload"], lang)
    return render_template("workload_teacher.html", teacher=dict(teacher),
                           entries=entries, total_credits=total_credits,
                           status=status)


# ════════════════════════════════════════════════════════════
# ОТЧЁТЫ
# ════════════════════════════════════════════════════════════
@app.route("/reports")
@login_required
def reports_page():
    lang     = get_current_lang()
    semester = request.args.get("semester", type=int)
    return render_template("reports.html",
                           all_teachers    = rep.report_all_teachers(lang),
                           overloaded      = rep.report_overloaded_teachers(lang),
                           underloaded     = rep.report_underloaded_teachers(lang=lang),
                           subjects_report = rep.report_subjects_summary(),
                           semester_data   = rep.report_by_semester(semester) if semester else [],
                           selected_sem    = semester,
                           semesters       = SEMESTERS)


# ════════════════════════════════════════════════════════════
# УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ (только admin)
# ════════════════════════════════════════════════════════════
@app.route("/users")
@login_required
@admin_required
def users_list():
    return render_template("users.html", users=au.get_all_users(),
                           roles=au.ROLES)


@app.route("/users/add", methods=["GET", "POST"])
@login_required
@admin_required
def user_add():
    if request.method == "POST":
        username  = request.form.get("username", "").strip()
        password  = request.form.get("password", "").strip()
        full_name = request.form.get("full_name", "").strip()
        role      = request.form.get("role", "viewer")
        if not username or not password or not full_name:
            flash(_t("users.flash.required", get_current_lang()), "danger")
        elif len(password) < 6:
            flash(_t("users.flash.short_pwd", get_current_lang()), "danger")
        else:
            try:
                au.create_user(username, password, full_name, role)
                flash(_t("users.flash.created", get_current_lang(), name=username), "success")
                return redirect(url_for("users_list"))
            except ValueError as e:
                flash(str(e), "danger")
    return render_template("user_form.html", action="add", user=None,
                           roles=au.ROLES)


@app.route("/users/edit/<int:uid>", methods=["GET", "POST"])
@login_required
@admin_required
def user_edit(uid):
    users  = au.get_all_users()
    user   = next((u for u in users if u["id"] == uid), None)
    if not user:
        flash(_t("users.flash.not_found", get_current_lang()), "danger")
        return redirect(url_for("users_list"))
    if request.method == "POST":
        full_name    = request.form.get("full_name", "").strip()
        role         = request.form.get("role", "viewer")
        is_active    = bool(request.form.get("is_active"))
        new_password = request.form.get("new_password", "").strip()
        if new_password and len(new_password) < 6:
            flash(_t("users.flash.short_pwd", get_current_lang()), "danger")
        else:
            try:
                au.update_user(uid, full_name, role, is_active, new_password)
                flash(_t("users.flash.updated", get_current_lang()), "success")
                return redirect(url_for("users_list"))
            except ValueError as e:
                flash(str(e), "danger")
    return render_template("user_form.html", action="edit", user=user,
                           roles=au.ROLES)


@app.route("/users/delete/<int:uid>", methods=["POST"])
@login_required
@admin_required
def user_delete(uid):
    if uid == current_user.id:
        flash(_t("users.flash.no_self", get_current_lang()), "danger")
        return redirect(url_for("users_list"))
    try:
        au.delete_user(uid)
        flash(_t("users.flash.deleted", get_current_lang()), "warning")
    except ValueError as e:
        flash(str(e), "danger")
    return redirect(url_for("users_list"))


# ── Профиль / смена пароля ───────────────────────────────────
@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    if request.method == "POST":
        old_pw  = request.form.get("old_password", "")
        new_pw  = request.form.get("new_password", "").strip()
        confirm = request.form.get("confirm_password", "").strip()
        from werkzeug.security import check_password_hash
        row = au.get_user_by_username(current_user.username)
        if not check_password_hash(row["password_hash"], old_pw):
            flash(_t("profile.flash.wrong_old", get_current_lang()), "danger")
        elif new_pw != confirm:
            flash(_t("profile.flash.mismatch", get_current_lang()), "danger")
        else:
            try:
                au.change_password(current_user.id, new_pw)
            except ValueError as e:
                flash(str(e), "danger")
            else:
                flash(_t("profile.flash.changed", get_current_lang()), "success")
                return redirect(url_for("index"))
    return render_template("profile.html")


# ════════════════════════════════════════════════════════════
# ИИ-АССИСТЕНТ
# ════════════════════════════════════════════════════════════

@app.route("/ai")
@login_required
def ai_page():
    dash = ai.get_dashboard()
    return render_template("ai.html", dash=dash)

@app.route("/ai/retrain", methods=["POST"])
@login_required
@editor_required
def ai_retrain():
    ai.retrain()
    flash(_t("ai.flash.retrained", get_current_lang()), "success")
    return redirect(url_for("ai_page"))

@app.route("/ai/find-teacher")
@login_required
def ai_find_teacher():
    credits_arg = request.args.get("credits", request.args.get("hours", "5"))
    try:
        credits = float(str(credits_arg).replace(",", "."))
    except ValueError:
        credits = 5.0
    results = ai.find_best_teacher(credits)
    return jsonify({"results": results})

@app.route("/ai/predict")
@login_required
def ai_predict():
    position    = request.args.get("position", "Доцент")
    credits_arg = request.args.get("credits", request.args.get("hours", "5"))
    try:
        credits = float(str(credits_arg).replace(",", "."))
    except ValueError:
        credits = 5.0
    result = ai.predict_for_position(position, credits)
    return jsonify(result)

@app.route("/ai/chat", methods=["POST"])
@login_required
def ai_chat():
    """Чат-бот на базе Groq API — отвечает на любые вопросы."""
    from config import GROQ_API_KEY
    from groq import Groq

    data     = request.get_json(silent=True) or {}
    user_msg = (data.get("message") or "").strip()
    history  = data.get("history") or []

    if not user_msg:
        return jsonify({"error": "Пустой запрос"}), 400

    if not GROQ_API_KEY:
        return jsonify({"reply": (
            "⚠️ API ключ Groq не настроен.\n\n"
            "Получите бесплатный ключ на https://console.groq.com\n"
            "Затем запустите: set GROQ_API_KEY=gsk_..."
        )})

    # Собираем актуальные данные кафедры для контекста
    dash     = ai.get_dashboard()
    teachers = dash["teachers"]
    anomalies = dash["anomalies"]
    summary  = dash["summary"]

    teacher_lines = []
    for t in teachers:
        st = workload_status(t["total_credits"], t["max_workload"])
        teacher_lines.append(
            f"- {t['full_name']} ({t['position']}): "
            f"{t['total_credits']}/{t['max_workload']} кр., {st['percent']}% — {st['label']}"
        )

    anomaly_lines = [f"- {a['teacher']}: {a['message']}" for a in anomalies] or ["- Проблем не обнаружено"]

    lang = get_current_lang()
    dept_local = _t("project.department", lang)
    lang_directive = {
        "kk": "Қазақ тілінде жауап бер.",
        "ru": "Отвечай на русском языке.",
    }.get(lang, "Отвечай на русском языке.")
    system_prompt = f"""Ты — умный ИИ-ассистент {dept_local} (тема проекта: автоматизация работы преподавателей).
{lang_directive} Будь дружелюбным, конкретным и полезным.
Отвечай на ЛЮБЫЕ вопросы — как по кафедре, так и на общие темы (наука, программирование, математика, жизнь и т.д.).
Нагрузка измеряется в кредитах: согласно П ЕНУ K.V.14-22 (Приложение 3) 1 кредит = 15 академических часов лекционных/практических/лабораторных занятий.

АКТУАЛЬНЫЕ ДАННЫЕ КАФЕДРЫ (обновляются в реальном времени):
Преподавателей: {summary['teachers']}, дисциплин: {summary['subjects']}
Всего кредитов: {summary['total_credits']}/{summary['total_max']} кр. ({summary['dept_pct']}% загруженность)
Баланс кафедры: {dash['balance_score']}/100

ПРЕПОДАВАТЕЛИ И НАГРУЗКА:
{chr(10).join(teacher_lines)}

ОБНАРУЖЕННЫЕ ПРОБЛЕМЫ:
{chr(10).join(anomaly_lines)}

Когда отвечаешь на вопросы о кафедре — используй эти данные. На общие вопросы отвечай из своих знаний."""

    # Строим историю сообщений
    messages = [{"role": "system", "content": system_prompt}]
    for m in history[-10:]:
        if m.get("role") in ("user", "assistant"):
            messages.append({"role": m["role"], "content": m["content"]})
    messages.append({"role": "user", "content": user_msg})

    try:
        client = Groq(api_key=GROQ_API_KEY)
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=messages,
            max_tokens=1024,
            temperature=0.7,
        )
        reply = response.choices[0].message.content
        return jsonify({"reply": reply})

    except Exception as e:
        err = str(e)
        if "401" in err or "invalid_api_key" in err.lower():
            return jsonify({"reply": "⚠️ Неверный API ключ Groq. Проверьте GROQ_API_KEY."})
        if "rate_limit" in err.lower():
            return jsonify({"reply": "⚠️ Превышен лимит запросов Groq. Попробуйте через минуту."})
        return jsonify({"reply": f"⚠️ Ошибка Groq API: {err[:200]}"})


# ════════════════════════════════════════════════════════════
# ЗАПУСК
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"\n{'='*55}")
    print(f"  {DEPARTMENT_NAME}")
    print(f"  {PROJECT_THEME}")
    print(f"  Система управления учебной нагрузкой")
    print(f"{'='*55}")
    print(f"  Откройте браузер: http://127.0.0.1:5000")
    if config.ALLOW_DEFAULT_ADMIN_PASSWORD:
        print(f"  Логин по умолчанию: admin / admin1234 (DEBUG)")
    else:
        print(f"  Пароль администратора см. в логах при первом запуске.")
    print(f"{'='*55}\n")
    # host=127.0.0.1 — в DEBUG не биндимся на 0.0.0.0 (безопасность).
    # Для внешнего доступа используйте gunicorn/uwsgi за reverse-proxy.
    host = "127.0.0.1" if DEBUG else "0.0.0.0"
    app.run(debug=DEBUG, host=host, port=5000)